import os
import pandas as pd
from openpyxl import load_workbook

script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

data = pd.read_spss('PODES2026-DESA.sav', convert_categoricals=False)

# daftar provinsi dari 2 digit pertama kode kabupaten
daftar_prov = sorted(data['r102'].astype(str).str[:2].unique())

for kode_prov in daftar_prov:

    print(f'Proses Provinsi {kode_prov}')

    # Filter hanya kabupaten/kota dalam provinsi tersebut
    data_prov = data[
        data['r102'].astype(str).str[:2] == kode_prov
    ].copy()

    wb = load_workbook('template.xlsx')

    # Pembuatan Tabel
    # Tabel 1.1
    tabel_1_1 = (
        data_prov.groupby('r102').agg(
            nama_kabupaten=('nama_kab', 'first'),
            jumlah_kecamatan=('r103', 'nunique'),
            jumlah_desa=('iddesa', 'nunique')
        ).reset_index(drop=True)
    )
    tabel_1_1.insert(tabel_1_1.columns.get_loc('nama_kabupaten') + 1, 'kosong', '')
    tabel_1_1.loc[len(tabel_1_1)] = [
        'PROVINSI', '',
        tabel_1_1['jumlah_kecamatan'].sum(),
        tabel_1_1['jumlah_desa'].sum()
    ]

    # Tabel 1.2
    tabel_1_2 = (
        data_prov.groupby('r102').agg(
            nama_kabupaten=('nama_kab', 'first'),
            desa=('r301', lambda x: x.isin([1, 4]).sum()),
            kelurahan=('r301', lambda x: (x == 2).sum()),
            upt=('r301', lambda x: (x == 3).sum()),
            total=('r301', lambda x: x.isin([1, 2, 3, 4]).sum())
        ).reset_index(drop=True)
    )
    tabel_1_2.insert(tabel_1_2.columns.get_loc('nama_kabupaten') + 1, 'kosong', '')
    tabel_1_2.loc[len(tabel_1_2)] = [
        'PROVINSI', '',
        tabel_1_2['desa'].sum(),
        tabel_1_2['kelurahan'].sum(),
        tabel_1_2['upt'].sum(),
        tabel_1_2['total'].sum()
    ]

    # Tabel 1.3
    tabel_1_3 = (
        data_prov.groupby('r102').agg(
            nama_kabupaten=('nama_kab', 'first'),
            dalam=('r303a', lambda x: (x == 1).sum()),
            tepi=('r303a', lambda x: (x == 2).sum()),
            luar=('r303a', lambda x: (x == 3).sum()),
            jumlah=('r303a', lambda x: x.isin([1, 2, 3]).sum())
        ).reset_index(drop=True)
    )
    tabel_1_3.insert(tabel_1_3.columns.get_loc('nama_kabupaten') + 1, 'kosong', '')
    tabel_1_3.loc[len(tabel_1_3)] = [
        'PROVINSI', '',
        tabel_1_3['dalam'].sum(),
        tabel_1_3['tepi'].sum(),
        tabel_1_3['luar'].sum(),
        tabel_1_3['jumlah'].sum()
    ]

    # Tabel 1.4
    tabel_1_4 = (
        data_prov.groupby('r102').agg(
            nama_kabupaten=('nama_kab', 'first'),
            tepi_laut=('r302a', lambda x: (x == 1).sum()),
            bukan=('r302a', lambda x: (x == 2).sum()),
            jumlah=('r302a', lambda x: x.isin([1, 2]).sum())
        ).reset_index(drop=True)
    )
    tabel_1_4.insert(tabel_1_4.columns.get_loc('nama_kabupaten') + 1,'kosong', '')
    tabel_1_4.loc[len(tabel_1_4)] = [
        'PROVINSI', '',
        tabel_1_4['tepi_laut'].sum(),
        tabel_1_4['bukan'].sum(),
        tabel_1_4['jumlah'].sum()
    ]

    # Tabel 1.5
    tabel_1_5 = (
        data_prov.groupby('r102').agg(
            nama_kabupaten=('nama_kab', 'first'),
            pertanian=('r1008a', lambda x: (x == 1).sum()),
            industri=('r1008a', lambda x: (x == 2).sum()),
            jasa=('r1008a', lambda x: (x == 3).sum()),
            jumlah=('r1008a', lambda x: x.isin([1, 2, 3]).sum())
        ).reset_index(drop=True)
    )
    tabel_1_5.insert(tabel_1_5.columns.get_loc('nama_kabupaten') + 1, 'kosong', '')
    tabel_1_5.loc[len(tabel_1_5)] = [
        'PROVINSI', '',
        tabel_1_5['pertanian'].sum(),
        tabel_1_5['industri'].sum(),
        tabel_1_5['jasa'].sum(),
        tabel_1_5['jumlah'].sum()
    ]

    # Tabel 1.6
    tabel_1_6 = (
        data_prov.groupby('r102').agg(
            nama_kabupaten=('nama_kab', 'first'),
            pangan=('r1008b', lambda x: (x == 1).sum()),
            horti=('r1008b', lambda x: (x == 2).sum()),
            kebun=('r1008b', lambda x: (x == 3).sum()),
            ternak=('r1008b', lambda x: (x == 4).sum()),
            ikan=('r1008b', lambda x: (x == 5).sum()),
            hutan=('r1008b', lambda x: (x == 6).sum()),
            jasa=('r1008b', lambda x: (x == 7).sum()),
            jumlah=('r1008b', lambda x: x.isin([1, 2, 3, 4, 5, 6, 7]).sum())
        ).reset_index(drop=True)
    )
    tabel_1_6.insert(tabel_1_6.columns.get_loc('nama_kabupaten') + 1, 'kosong', '')
    tabel_1_6.loc[len(tabel_1_6)] = [
        'PROVINSI', '',
        tabel_1_6['pangan'].sum(),
        tabel_1_6['horti'].sum(),
        tabel_1_6['kebun'].sum(),
        tabel_1_6['ternak'].sum(),
        tabel_1_6['ikan'].sum(),
        tabel_1_6['hutan'].sum(),
        tabel_1_6['jasa'].sum(),
        tabel_1_6['jumlah'].sum(),
    ]

    # Tabel 2.1
    tabel_2_1 = (
        data_prov.groupby('r102').agg(
            nama_kabupaten=('nama_kab', 'first'),
            pln=('r501a1', lambda x: (x > 0).sum()),
            non_pln=('r501a2', lambda x: (x > 0).sum()),
            non_listrik=('r501b', lambda x: (x > 0).sum()),
        ).reset_index(drop=True)
    )
    tabel_2_1.insert(tabel_2_1.columns.get_loc('nama_kabupaten') + 1, 'kosong', '')
    tabel_2_1.loc[len(tabel_2_1)] = [
        'PROVINSI', '',
        tabel_2_1['pln'].sum(),
        tabel_2_1['non_pln'].sum(),
        tabel_2_1['non_listrik'].sum(),
    ]

    # Tabel 2.2
    tabel_2_2 = (
        data_prov.groupby('r102').agg(
            nama_kabupaten=('nama_kab', 'first'),
            listrik_pemerintah=('r502b', lambda x: (x == 1).sum()),
            listrik_non_pemerintah=('r502b', lambda x: (x == 2).sum()),
            non_listrik=('r502b', lambda x: (x == 3).sum()),
            tidak_ada=('r502a', lambda x: (x == 3).sum()),
            jumlah=('r502a', lambda x: x.isin([1, 2, 3]).sum()),
        ).reset_index(drop=True)
    )
    tabel_2_2.insert(tabel_2_2.columns.get_loc('nama_kabupaten') + 1, 'kosong', '')
    tabel_2_2.loc[len(tabel_2_2)] = [
        'PROVINSI', '',
        tabel_2_2['listrik_pemerintah'].sum(),
        tabel_2_2['listrik_non_pemerintah'].sum(),
        tabel_2_2['non_listrik'].sum(),
        tabel_2_2['tidak_ada'].sum(),
        tabel_2_2['jumlah'].sum(),
    ]

    # Tabel 2.3
    tabel_2_3 = (
        data_prov.groupby('r102').agg(
            nama_kabupaten=('nama_kab', 'first'),
            listrik=('r503', lambda x: (x > 0).sum()),
            lpg_55=('r503', lambda x: (x > 0).sum()),
            lpg_12=('r503', lambda x: (x > 0).sum()),
            lpg_3=('r503', lambda x: (x > 0).sum()),
            gas_kota=('r503', lambda x: (x > 0).sum()),
            biogas=('r503', lambda x: (x > 0).sum()),
            minyak_tanah=('r503', lambda x: (x > 0).sum()),
            briket=('r503', lambda x: (x > 0).sum()),
            arang=('r503', lambda x: (x > 0).sum()),
            kayu_bakar=('r503', lambda x: (x > 0).sum()),
            lainnya=('r503', lambda x: (x > 0).sum()),
            jumlah=('r503', lambda x: x.isin([1,2,3,4,5,6,7,8,9,10,11]).sum()),
        ).reset_index(drop=True)
    )
    tabel_2_3.insert(tabel_2_3.columns.get_loc('nama_kabupaten') + 1, 'kosong', '')
    tabel_2_3.loc[len(tabel_2_3)] = [
        'PROVINSI', '',
        tabel_2_3['listrik'].sum(),
        tabel_2_3['lpg_55'].sum(),
        tabel_2_3['lpg_12'].sum(),
        tabel_2_3['lpg_3'].sum(),
        tabel_2_3['gas_kota'].sum(),
        tabel_2_3['biogas'].sum(),
        tabel_2_3['minyak_tanah'].sum(),
        tabel_2_3['briket'].sum(),
        tabel_2_3['arang'].sum(),
        tabel_2_3['kayu_bakar'].sum(),
        tabel_2_3['lainnya'].sum(),
        tabel_2_3['jumlah'].sum(),
    ]

    # Tabel 2.4
    tabel_2_4 = (
        data_prov.groupby('r102').agg(
            nama_kabupaten=('nama_kab', 'first'),
            agen_minyak=('r1001a', lambda x: (x == 1).sum()),
            agen_lpg=('r1001b', lambda x: (x == 1).sum()),
        ).reset_index(drop=False)
    )
    tidak_ada = (
    data_prov.assign(
        tidak_ada=((data_prov['r1001a'] == 2) &
                   (data_prov['r1001b'] == 2))
    ).groupby('r102')['tidak_ada'].sum().reset_index()
    )
    tabel_2_4 = tabel_2_4.merge(tidak_ada, on='r102')
    tabel_2_4.insert(tabel_2_4.columns.get_loc('nama_kabupaten') + 1, 'kosong', '')
    tabel_2_4 = tabel_2_4.drop(columns=['r102'])
    tabel_2_4.loc[len(tabel_2_4)] = [
        'PROVINSI', '',
        tabel_2_4['agen_minyak'].sum(), 
        tabel_2_4['agen_lpg'].sum(), 
        tabel_2_4['tidak_ada'].sum(),
    ]



    # Tabel 4.1
    tabel_4_1 = (
        data_prov.groupby('r102').agg(
            nama_kabupaten=('nama_kab', 'first'),
            darat=('r901a', lambda x: (x == 1).sum()),
            air=('r901a', lambda x: (x == 2).sum()),
            darat_air=('r901a', lambda x: (x == 3).sum()),
            udara=('r901a', lambda x: (x == 4).sum()),
        ).reset_index(drop=True)
    )
    tabel_4_1.insert(tabel_4_1.columns.get_loc('nama_kabupaten') + 1, 'kosong', '')
    tabel_4_1.loc[len(tabel_4_1)] = [
        'PROVINSI', '',
        tabel_4_1['darat'].sum(),
        tabel_4_1['air'].sum(),
        tabel_4_1['darat_air'].sum(),
        tabel_4_1['udara'].sum(),
    ]

    # Tabel 4.2
    tabel_4_2 = (
        data_prov.groupby('r102').agg(
            nama_kabupaten=('nama_kab', 'first'),
            trayek_tetap_non_trayek=('r901c1', lambda x: (x == 'AB').sum()),
            trayek_tetap=('r901c1', lambda x: (x == 'A').sum()),
            non_trayek=('r901c1', lambda x: (x == 'B').sum()),
            tidak_ada=('r901c1', lambda x: (x == 'X').sum()),
            jumlah=('r901c1', lambda x: x.isin(['A', 'B', 'AB', 'X']).sum()),
        ).reset_index(drop=True)
    )
    tabel_4_2.insert(tabel_4_2.columns.get_loc('nama_kabupaten') + 1, 'kosong', '')
    tabel_4_2.loc[len(tabel_4_2)] = [
        'PROVINSI', '',
        tabel_4_2['trayek_tetap_non_trayek'].sum(),
        tabel_4_2['trayek_tetap'].sum(),
        tabel_4_2['non_trayek'].sum(),
        tabel_4_2['tidak_ada'].sum(),
        tabel_4_2['jumlah'].sum(),
    ]

    # Mulai memproduksi tabel (file)
    # Isi semua sheet sesuai nama tabel
    for sheet_name in wb.sheetnames:

        tabel = f"tabel_{sheet_name.replace('.', '_')}"

        if tabel not in globals():
            print(f'{tabel} tidak ditemukan')
            continue

        df_sheet = globals()[tabel]

        ws = wb[sheet_name]

        for r, row in enumerate(df_sheet.values, start=10):
            for c, value in enumerate(row, start=1):
                ws.cell(r, c, value)

    # Simpan file provinsi
    os.makedirs(kode_prov, exist_ok=True)

    wb.save(f'{kode_prov}/{kode_prov}00.xlsx')
